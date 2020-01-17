# -*- coding: utf-8 -*
import datetime
import sys
import glob, os
import os.path
import sqlite3
import platform
from openpyxl import load_workbook
import shutil
import winsound

def main():
    LerArquivoDiretorio()

def LerArquivoDiretorio():
    try:
        dt = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"); print ("{} - Processando ...".format(dt))

        os.chdir("rmspendentes")
        for i in os.listdir(os.getcwd()):
            if os.path.isfile(i):
                if (i.endswith(".xls") or i.endswith(".xlsx")):
                    NomeDoArquivo = (os.path.join(os.getcwd(), i))
                    NomeDoArquivo1 = (os.path.basename(NomeDoArquivo))
                    LerPlanilha (NomeDoArquivo)
        dt = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"); print ("{} - Fim ...".format(dt))
    except Exception as e:
        print ("===========> Ocorreu um erro na rotina LerArquivoDiretorio")
        raise

def LerPlanilha(NomeDoArquivo):
    NomeDaplanilha = "Calendario de Mudanças Geral"
    TotalDeLinhas = 0

    try:
        wb = load_workbook(NomeDoArquivo)
        ws_rms = wb[NomeDaplanilha]
        TotalDeLinhas = ws_rms.max_row
        print("Total de Linhas na aba {}: {}".format(NomeDaplanilha,TotalDeLinhas))
        print("Lendo Planilha ...")
        LerLinhasDaPlanilha(TotalDeLinhas, ws_rms)
            
    except Exception as e:
        print ("===========> Ocorreu um erro na rotina LerPlanilha, no load workbook")
        raise

def LerLinhasDaPlanilha(TotalDeLinhas, ws_rms):
    LinhaInicial = 5
    
    ColNumeroRm = 2
    ColResumo_Descricao_RM = 3
    colResponsavel = 4
    colData_criacao = 6
    colTipoDaMudanca = 7
    colDiretoria = 13
    colSistema = 14
    colOrigem = 17
    colDataIniDesenv = 18
    colDataFimDesenv = 19
    colDataIniQA = 20
    colDataFimQA = 21
    colDataIniHML = 22
    colDataFimHML = 23
    colDataIniPrd = 24
    colDataFimPrd = 25
    colStatus = 26
    colDataComite = 29
    colMotivo = 30
    colCenarioNegAtual = 31
    colCenarioNegDesejado = 32
    colCenarioTecProposto = 33
    TicketdeSistemaExterno = 34

    listarms = {}

    try:
        for linha in range(LinhaInicial,TotalDeLinhas+1):
            listarms.clear()
            #print("Número da RM: {}".format(ws_rms.cell(row=linha, column=ColNumeroRm).value))
            listarms ['ColNumeroRm'] = ws_rms.cell(row=linha, column=ColNumeroRm).value
            listarms ['ColResumo_Descricao_RM'] = ws_rms.cell(row=linha, column=ColResumo_Descricao_RM).value
            listarms ['colResponsavel'] = ws_rms.cell(row=linha, column=colResponsavel).value
            listarms ['colData_criacao'] = ws_rms.cell(row=linha, column=colData_criacao).value
            listarms ['colTipoDaMudanca'] = ws_rms.cell(row=linha, column=colTipoDaMudanca).value
            listarms ['colDiretoria'] = ws_rms.cell(row=linha, column=colDiretoria).value
            listarms ['colSistema'] = ws_rms.cell(row=linha, column=colSistema).value
            listarms ['colOrigem'] = ws_rms.cell(row=linha, column=colOrigem).value
            listarms ['colDataIniDesenv'] = ws_rms.cell(row=linha, column=colDataIniDesenv).value
            listarms ['colDataFimDesenv'] = ws_rms.cell(row=linha, column=colDataFimDesenv).value
            listarms ['colDataIniQA'] = ws_rms.cell(row=linha, column=colDataIniQA).value
            listarms ['colDataFimQA'] = ws_rms.cell(row=linha, column=colDataFimQA).value
            listarms ['colDataIniHML'] = ws_rms.cell(row=linha, column=colDataIniHML).value
            listarms ['colDataFimHML'] = ws_rms.cell(row=linha, column=colDataFimHML).value
            listarms ['colDataIniPrd'] = ws_rms.cell(row=linha, column=colDataIniPrd).value
            listarms ['colDataFimPrd'] = ws_rms.cell(row=linha, column=colDataFimPrd).value
            listarms ['colStatus'] = ws_rms.cell(row=linha, column=colStatus).value
            listarms ['colDataComite'] = ws_rms.cell(row=linha, column=colDataComite).value
            listarms ['colMotivo'] = ws_rms.cell(row=linha, column=colMotivo).value
            listarms ['colCenarioNegAtual'] = ws_rms.cell(row=linha, column=colCenarioNegAtual).value
            listarms ['colCenarioNegDesejado'] = ws_rms.cell(row=linha, column=colCenarioNegDesejado).value
            listarms ['colCenarioTecProposto'] = ws_rms.cell(row=linha, column=colCenarioTecProposto).value
            listarms ['colTicketdeSistemaExterno'] = ws_rms.cell(row=linha, column=TicketdeSistemaExterno).value
            insert_rms(listarms)
        print ("fim")
    except Exception as e:
        print ("===========> Ocorreu um erro na rotina LerLinhasDaPlanilha, na linha:{}".format(linha))
        raise

def insert_rms(listarms):
    try:
        BASE_DIR = os.path.dirname(os.path.abspath(__file__))
        db_path = os.path.join(BASE_DIR, "Projetos.db")
        
        conn = sqlite3.connect(db_path)
        cursor = conn.cursor()

        cursor.execute("""
        INSERT INTO rms (Numero_RM, Resumo_Descricao_RM, Responsavel, Data_criacao, TipoDaMudanca,
                              Diretoria, Sistema, Origem, DataIniDesenv, DataFimDesenv,
                              DataIniQA, DataFimQA, DataIniHML, DataFimHML, DataIniPrd,
                              DataFimPrd, Status, DataComite, Motivo, CenarioNegAtual,
                              CenarioNegDesejado, CenarioTecProposto, TicketdeSistemaExterno)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """, (listarms ['ColNumeroRm'], listarms ['ColResumo_Descricao_RM'], listarms ['colResponsavel'], listarms ['colData_criacao'], listarms ['colTipoDaMudanca'],
              listarms ['colDiretoria'],listarms ['colSistema'], listarms ['colOrigem'], listarms ['colDataIniDesenv'], listarms ['colDataFimDesenv'],
              listarms ['colDataIniQA'],listarms ['colDataFimQA'], listarms ['colDataIniHML'], listarms ['colDataFimHML'], listarms ['colDataIniPrd'],
              listarms ['colDataFimPrd'], listarms ['colStatus'],  listarms ['colDataComite'], listarms ['colMotivo'],     listarms ['colCenarioNegAtual'],
              listarms ['colCenarioNegDesejado'], listarms ['colCenarioTecProposto'],listarms ['colTicketdeSistemaExterno'],))

        conn.commit()

    except Exception as e:
        raise
    finally:
        conn.close()
                  
              
if __name__ == "__main__":
	main()
