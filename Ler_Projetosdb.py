# -*- coding: utf-8 -*
import datetime
import sys
import glob, os
import os.path
import sqlite3
import platform
from openpyxl import load_workbook

def main():
    lerdir01()
    #lerdir03()

def lerdir01():
    colNome_do_Projeto = 1
    colVP = 3
    colFormula_Fase = 95
    colA_N = 45
    colGP = 46
    colLT = 47
    colLider_de_Testes = 48
    colGerente_de_Desenvolvimento = 49
    colFormula_Status_Projeto = 107
    colDescricao = 44
    colInicio_desenv = 29
    colTermino_Desenv = 30
    colComplitude1 = 31
    colInicio_Testes_Integrados = 32
    colTermino_Testes_Integrados = 33
    colComplitude2 = 34
    colInicio_da_HML = 35
    colTermino_da_HML = 36
    colComplitude3 = 37
    colProblemas_Riscos = 62
    colPlano_de_Acao = 63
    colCausa = 64
    colSubCausa_Capacity = 65
    
    colLeonardo = "Leonardo Augusto Mendes Leandro"
    colJuliana = "Juliana Alves Castro Perez"
    
    os.chdir("pendentes")
    #for file in glob.glob("*.xlsm"):
    #print(file)
    projetos = {}
    for i in os.listdir(os.getcwd()):
        #print(i)
        projetos.clear();
        print (os.path.join(os.getcwd(), i))
        if os.path.isfile(i):
            if i.endswith(".xls") or i.endswith(".xlsx") or i.endswith(".xlsm"):

                DataCriacaoDoArquivo = creation_date(os.path.join(os.getcwd(), i))
                print ("Data criação do Arquivo: {}".format(datetime.datetime.fromtimestamp(DataCriacaoDoArquivo)))


                NomeDoArquivo = (os.path.join(os.getcwd(), i))
                wb = load_workbook(os.path.join(os.getcwd(), i))
                ws_Projetos = wb['Projetos']
                total_linhas = ws_Projetos.max_row
                print("Total de Linhas na aba projetos: {}".format(total_linhas))

                id_arquivo = Insert_tb_arquivos(NomeDoArquivo, DataCriacaoDoArquivo, total_linhas)
                print ("Last id: {}".format(id_arquivo))
                
                linha = 4
                projetos ['id_arquivo'] = id_arquivo
                
                projetos ['Nome_do_Projeto'] = ws_Projetos.cell(row=linha, column=colNome_do_Projeto).value
                projetos ['VP'] = ws_Projetos.cell(row=linha, column=colVP).value
                projetos ['Formula_Fase'] = ws_Projetos.cell(row=linha, column=colFormula_Fase).value
                projetos ['A_N'] = ws_Projetos.cell(row=linha, column=colA_N).value
                projetos ['VP'] = ws_Projetos.cell(row=linha, column=colVP).value
                projetos ['GP'] = ws_Projetos.cell(row=linha, column=colGP).value
                projetos ['LT'] = ws_Projetos.cell(row=linha, column=colLT).value
                projetos ['Lider_de_Testes'] = ws_Projetos.cell(row=linha, column=colLider_de_Testes).value
                projetos ['Gerente_de_Desenvolvimento'] = ws_Projetos.cell(row=linha, column=colGerente_de_Desenvolvimento).value
                projetos ['Formula_Status_Projeto'] = ws_Projetos.cell(row=linha, column=colFormula_Status_Projeto).value
                projetos ['Descricao'] = ws_Projetos.cell(row=linha, column=colDescricao).value
                projetos ['Inicio_desenv'] = ws_Projetos.cell(row=linha, column=colInicio_desenv).value
                projetos ['Termino_Desenv'] = ws_Projetos.cell(row=linha, column=colTermino_Desenv).value
                projetos ['Complitude1'] = ws_Projetos.cell(row=linha, column=colComplitude1).value
                projetos ['Inicio_Testes_Integrados'] = ws_Projetos.cell(row=linha, column=colInicio_Testes_Integrados).value
                projetos ['Termino_Testes_Integrados'] = ws_Projetos.cell(row=linha, column=colTermino_Testes_Integrados).value
                projetos ['Complitude2'] = ws_Projetos.cell(row=linha, column=colComplitude2).value
                projetos ['Inicio_da_HML'] = ws_Projetos.cell(row=linha, column=colInicio_da_HML).value
                projetos ['Termino_da_HML'] = ws_Projetos.cell(row=linha, column=colTermino_da_HML).value
                projetos ['Complitude3'] = ws_Projetos.cell(row=linha, column=colComplitude3).value
                projetos ['Problemas_Riscos'] = ws_Projetos.cell(row=linha, column=colProblemas_Riscos).value
                projetos ['Plano_de_Acao'] = ws_Projetos.cell(row=linha, column=colPlano_de_Acao).value
                projetos ['Causa'] = ws_Projetos.cell(row=linha, column=colCausa).value
                projetos ['SubCausa_Capacity'] = ws_Projetos.cell(row=linha, column=colSubCausa_Capacity).value
                projetos ['Causa'] = ws_Projetos.cell(row=linha, column=colCausa).value

                #LerProjetos(**projetos)
                insert_new_proj(**projetos)

def Insert_tb_arquivos(NomeDoArquivo, DataCriacaoDoArquivo, NumeroDeLinhas):
    dt = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"); print ("{} - Conectando ao banco e inserindo linha ...".format(dt))
    lastid = 1
    
    strNomeDoArquivo = NomeDoArquivo
    dtDataCriacaoDoArquivo = datetime.datetime.fromtimestamp(DataCriacaoDoArquivo)
    intNumeroDeLinhas = NumeroDeLinhas

    BASE_DIR = os.path.dirname(os.path.abspath(__file__))
    db_path = os.path.join(BASE_DIR, "Projetos.db")
    

    conn = sqlite3.connect(db_path)
    cursor = conn.cursor()
    
##    # obtendo informações da tabela
##    nome_tabela = 'arquivos'
##    # obtendo o schema da tabela
##    cursor.execute("""
##    SELECT sql FROM sqlite_master WHERE type='table' AND name=?
##    """, (nome_tabela,))
##
##    print('Schema:')
##    for schema in cursor.fetchall():
##        print("%s" % (schema))
    

    cursor.execute("""
    INSERT INTO arquivos (NomeDoArquivo, DataCriacaoDoArquivo, TotalDeLinhas)
    VALUES (?, ?, ?)
    """, (strNomeDoArquivo, dtDataCriacaoDoArquivo,intNumeroDeLinhas))

    # gravando no bd
    conn.commit()

    dt = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"); print ("{} - DADOS INSERIDOS COM SUCESSO ...".format(dt))

    conn.close()

    lastid = cursor.lastrowid

    return lastid

def LerProjetos(**projetos):
    dt = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"); print ("{} - MONTANDO LISTA ... ...".format(dt))
   

    insert_db(linha)

def insert_new_proj(**projetos):
    dt = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"); print ("{} - Conectando ao banco e inserindo linha ...".format(dt))
    

    BASE_DIR = os.path.dirname(os.path.abspath(__file__))
    db_path = os.path.join(BASE_DIR, "Projetos.db")
    
    conn = sqlite3.connect(db_path)
    cursor = conn.cursor()
    
    linha = {}

    linha['Nome_do_Projeto']= projetos['Nome_do_Projeto']
    linha['VP']=projetos['VP']
    linha['Formula_Fase']=projetos['Formula_Fase']
    linha['A_N']=projetos['A_N']
    linha['GP']=projetos['GP']
    linha['LT']=projetos['LT']
    linha['Lider_de_Testes']=projetos ['Lider_de_Testes']
    linha['Gerente_de_Desenvolvimento']=projetos['Gerente_de_Desenvolvimento']
    linha['Formula_Status_Projeto']=projetos ['Formula_Status_Projeto']
    linha['Descricao']=projetos ['Descricao']
    linha['Nome_Arquivo']=""
    linha['Inicio_desenv']=projetos ['Inicio_desenv']
    linha['Termino_Desenv']=projetos ['Termino_Desenv']
    linha['Complitude1']=projetos ['Complitude1']
    linha['Inicio_Testes_Integrados']=projetos ['Inicio_Testes_Integrados']
    linha['Termino_Testes_Integrados']=projetos ['Termino_Testes_Integrados']
    linha['Complitude2']=projetos ['Complitude2']
    linha['Inicio_da_HML']=projetos ['Inicio_da_HML']
    linha['Termino_da_HML']=projetos ['Termino_da_HML']
    linha['Complitude3']=projetos ['Complitude3']
    linha['Problemas_Riscos']=projetos ['Problemas_Riscos']
    linha['SubCausa_Capacity']=projetos ['SubCausa_Capacity']
    linha['Plano_de_Acao']=projetos ['Plano_de_Acao']
    linha['Causa']=projetos ['Causa']
    linha['id_arquivo'] = projetos ['id_arquivo']

    Nome_do_Projeto = projetos['Nome_do_Projeto']
    VP = projetos['VP']
    Formula_Fase= projetos['Formula_Fase']
    A_N= projetos['A_N']
    GP = projetos['GP']
    LT = projetos['LT']
    Lider_de_Testes = projetos ['Lider_de_Testes']
    Gerente_de_Desenvolvimento = projetos['Gerente_de_Desenvolvimento']
    Formula_Status_Projeto = projetos ['Formula_Status_Projeto']
    Descricao = projetos ['Descricao']
    Nome_Arquivo = ""
    Inicio_desenv = projetos ['Inicio_desenv']
    Termino_Desenv = projetos ['Termino_Desenv']
    Complitude1 = projetos ['Complitude1']
    Inicio_Testes_Integrados = projetos ['Inicio_Testes_Integrados']
    Termino_Testes_Integrados = projetos ['Termino_Testes_Integrados']
    Complitude2 = projetos ['Complitude2']
    Inicio_da_HML = projetos ['Inicio_da_HML']
    Termino_da_HML = projetos ['Termino_da_HML']
    Complitude3 = projetos ['Complitude3']
    Problemas_Riscos = projetos ['Problemas_Riscos']
    SubCausa_Capacity = projetos ['SubCausa_Capacity']
    Plano_de_Acao = projetos ['Plano_de_Acao']
    Causa = projetos ['Causa']
    id_arquivo = projetos ['id_arquivo']
  
##    cursor.executemany("""
##    INSERT INTO projetos (Nome_Projeto,VP,Formula_Fase, an, gp, LT, Lider_Teste, Gerente_Desenv, Formula_Status_Projeto, Descricao, Nome_Arquivo, Ini_desenv, Term_desenv, Completude1, Ini_Teste_Integrado, Term_Teste_Integrado, Completude2,Ini_hml, Fim_hml, Completude3, Problema_Risco, SubCausa, Plano_Acao, causa, id_arquivo)
##    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?,?, ?, ?, ?, ?, ?, ?, ?)
##    """, linha)

    cursor.execute("""
    INSERT INTO projetos (Nome_Projeto,VP,Formula_Fase, an, gp, LT, Lider_Teste, Gerente_Desenv, Formula_Status_Projeto, Descricao, Nome_Arquivo, Ini_desenv, Term_desenv, Completude1, Ini_Teste_Integrado, Term_Teste_Integrado, Completude2,Ini_hml, Fim_hml, Completude3, Problema_Risco, SubCausa, Plano_Acao, causa, id_arquivo)
    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?,?, ?, ?, ?, ?, ?, ?, ?)
    """, (Nome_do_Projeto, VP, Formula_Fase, A_N, GP, LT, Lider_de_Testes, Gerente_de_Desenvolvimento, Formula_Status_Projeto, Descricao, Nome_Arquivo, Inicio_desenv, Termino_Desenv, Complitude1, Inicio_Testes_Integrados, Termino_Testes_Integrados, Complitude2, Inicio_da_HML, Termino_da_HML, Complitude3, Problemas_Riscos, SubCausa_Capacity, Plano_de_Acao, Causa, id_arquivo,))

    conn.commit()

    dt = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"); print ("{} - DADOS INSERIDOS COM SUCESSO ...".format(dt))

    conn.close()

def lerdir03():
    os.chdir("pendentes")
    files = glob.glob("*.xlsm")
    files.sort(key=os.path.getctime)
    #files.sort(key=os.path.getmtime)
    print("\n".join(files))
    files1 = ("\n".join(files))
    for file in files1:
        print(file)
        

def lerdir04():
    for root, dirs, files in os.walk("pendentes"):
        for file in files:
            if file.endswith(".xlsm"):
                 print(os.path.join(root, file))

def ler_arquivo(wb):
    _wb = wb
    dt = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    print("() - {}".format(dt,_wb.sheetnames))

def creation_date(path_to_file):
    """
    Try to get the date that a file was created, falling back to when it was
    last modified if that isn't possible.
    See http://stackoverflow.com/a/39501288/1709587 for explanation.
    """
    if platform.system() == 'Windows':
        return os.path.getctime(path_to_file)
    else:
        stat = os.stat(path_to_file)
        try:
            return stat.st_birthtime
        except AttributeError:
            # We're probably on Linux. No easy way to get creation dates here,
            # so we'll settle for when its content was last modified.
            return stat.st_mtime
if __name__ == "__main__":
	main()

