# -*- coding: utf-8 -*
import datetime
import sys
import glob, os
import os.path
import sqlite3
import platform
from openpyxl import load_workbook
import shutil
import winsound

def main():
    lerdir01()

def lerdir01():
    colNome_do_Projeto = 1
    colVP = 3
    colFormula_Fase = 95
    colA_N = 45
    colGP = 46
    colLT = 47
    colLider_de_Testes = 48
    colGerente_de_Desenvolvimento = 49
    colFormula_Status_Projeto = 107
    colDescricao = 44
    colInicio_desenv = 29
    colTermino_Desenv = 30
    colComplitude1 = 31
    colInicio_Testes_Integrados = 32
    colTermino_Testes_Integrados = 33
    colComplitude2 = 34
    colInicio_da_HML = 35
    colTermino_da_HML = 36
    colComplitude3 = 37
    colProblemas_Riscos = 62
    colPlano_de_Acao = 63
    colCausa = 64
    colSubCausa_Capacity = 65
    
    colLeonardo = "Leonardo Augusto Mendes Leandro"
    colJuliana = "Juliana Alves Castro Perez"
    
    os.chdir("pendentes")
    #for file in glob.glob("*.xlsm"):
    #print(file)
    projetos = {}
    for i in os.listdir(os.getcwd()):
        #print(i)
        projetos.clear();
        print (os.path.join(os.getcwd(), i))
        if os.path.isfile(i):
            if i.endswith(".xls") or i.endswith(".xlsx") or i.endswith(".xlsm"):

                DataCriacaoDoArquivo = creation_date(os.path.join(os.getcwd(), i))
                print ("Data criação do Arquivo: {}".format(datetime.datetime.fromtimestamp(DataCriacaoDoArquivo)))


                NomeDoArquivo = (os.path.join(os.getcwd(), i))

                NomeDoArquivo1 = (os.path.basename(NomeDoArquivo))
                
                wb = load_workbook(os.path.join(os.getcwd(), i))
                ws_Projetos = wb['Projetos']
                total_linhas = ws_Projetos.max_row
                print("Total de Linhas na aba projetos: {}".format(total_linhas))

                id_arquivo = Insert_tb_arquivos(NomeDoArquivo1, DataCriacaoDoArquivo, total_linhas)
                print ("Last id: {}".format(id_arquivo))
                
                linhainicial = 4

                for linha in range(linhainicial,total_linhas+1):
                    projetos ['id_arquivo'] = id_arquivo
                    projetos ['Nome_do_Projeto'] = ws_Projetos.cell(row=linha, column=colNome_do_Projeto).value
                    projetos ['VP'] = ws_Projetos.cell(row=linha, column=colVP).value
                    projetos ['Formula_Fase'] = ws_Projetos.cell(row=linha, column=colFormula_Fase).value
                    projetos ['A_N'] = ws_Projetos.cell(row=linha, column=colA_N).value
                    projetos ['VP'] = ws_Projetos.cell(row=linha, column=colVP).value
                    projetos ['GP'] = ws_Projetos.cell(row=linha, column=colGP).value
                    projetos ['LT'] = ws_Projetos.cell(row=linha, column=colLT).value
                    projetos ['Lider_de_Testes'] = ws_Projetos.cell(row=linha, column=colLider_de_Testes).value
                    projetos ['Gerente_de_Desenvolvimento'] = ws_Projetos.cell(row=linha, column=colGerente_de_Desenvolvimento).value
                    projetos ['Formula_Status_Projeto'] = ws_Projetos.cell(row=linha, column=colFormula_Status_Projeto).value
                    projetos ['Descricao'] = ws_Projetos.cell(row=linha, column=colDescricao).value
                    projetos ['Inicio_desenv'] = ws_Projetos.cell(row=linha, column=colInicio_desenv).value
                    projetos ['Termino_Desenv'] = ws_Projetos.cell(row=linha, column=colTermino_Desenv).value
                    projetos ['Complitude1'] = ws_Projetos.cell(row=linha, column=colComplitude1).value
                    projetos ['Inicio_Testes_Integrados'] = ws_Projetos.cell(row=linha, column=colInicio_Testes_Integrados).value
                    projetos ['Termino_Testes_Integrados'] = ws_Projetos.cell(row=linha, column=colTermino_Testes_Integrados).value
                    projetos ['Complitude2'] = ws_Projetos.cell(row=linha, column=colComplitude2).value
                    projetos ['Inicio_da_HML'] = ws_Projetos.cell(row=linha, column=colInicio_da_HML).value
                    projetos ['Termino_da_HML'] = ws_Projetos.cell(row=linha, column=colTermino_da_HML).value
                    projetos ['Complitude3'] = ws_Projetos.cell(row=linha, column=colComplitude3).value
                    projetos ['Problemas_Riscos'] = ws_Projetos.cell(row=linha, column=colProblemas_Riscos).value
                    projetos ['Plano_de_Acao'] = ws_Projetos.cell(row=linha, column=colPlano_de_Acao).value
                    projetos ['Causa'] = ws_Projetos.cell(row=linha, column=colCausa).value
                    projetos ['SubCausa_Capacity'] = ws_Projetos.cell(row=linha, column=colSubCausa_Capacity).value
                    projetos ['Causa'] = ws_Projetos.cell(row=linha, column=colCausa).value
                    Analisa_Para_Insert(**projetos)
                    
                #end for
                #Move o arquivo lido para o diretórios de lidos
                origem = NomeDoArquivo
                destino = origem.replace("pendentes","lidos")
                shutil.move (origem, destino)
                print ("Arquivo movido com sucesso!")
                Freq = 2500 # Set Frequency To 2500 Hertz
                Dur = 1000 # Set Duration To 1000 ms == 1 second
                winsound.Beep(Freq,Dur)

def Insert_tb_arquivos(NomeDoArquivo, DataCriacaoDoArquivo, NumeroDeLinhas):
    try:
        dt = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"); print ("{} - Conectando ao banco e inserindo linha ...".format(dt))
        lastid = 1
        
        strNomeDoArquivo = NomeDoArquivo
        dtDataCriacaoDoArquivo = datetime.datetime.fromtimestamp(DataCriacaoDoArquivo)
        intNumeroDeLinhas = NumeroDeLinhas

        BASE_DIR = os.path.dirname(os.path.abspath(__file__))
        db_path = os.path.join(BASE_DIR, "Projetos.db")
        

        conn = sqlite3.connect(db_path)
        cursor = conn.cursor()
       

        cursor.execute("""
        INSERT INTO arquivos (NomeDoArquivo, DataCriacaoDoArquivo, TotalDeLinhas)
        VALUES (?, ?, ?)
        """, (strNomeDoArquivo, dtDataCriacaoDoArquivo,intNumeroDeLinhas))

        # gravando no bd
        conn.commit()

        dt = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"); print ("{} - DADOS INSERIDOS COM SUCESSO ...".format(dt))

        conn.close()

        lastid = cursor.lastrowid

        return lastid
    except Exception as e:
        print ("Erro gerado: \n {}".format(str(e)))
        raise

def Analisa_Para_Insert(**projetos):
    try:
        dt = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"); print ("{} - Analise para insert...".format(dt))

        BASE_DIR = os.path.dirname(os.path.abspath(__file__))
        db_path = os.path.join(BASE_DIR, "Projetos.db")
        
        conn = sqlite3.connect(db_path)
        cursor = conn.cursor()
        
        Nome_do_Projeto = projetos['Nome_do_Projeto']

        if len(str(Nome_do_Projeto)) > 5:
            print("Analisando Projeto: {}".format(Nome_do_Projeto))

            cursor.execute("""
            SELECT count(*) tot FROM projetos
            WHERE Nome_Projeto = ?
            """, (Nome_do_Projeto, ))

            total_fetchone = cursor.fetchone()[0]
            
            print("Total Encontrado: {}".format(total_fetchone))
            
            if total_fetchone == 0:
                insert_new_proj(**projetos)
            else:
                Analise_Alteracoes(**projetos)
        else:
            print ("===========> Linha em branco")
            
        conn.close()
    except Exception as e:
        print ("Erro gerado: \n {}".format(str(e)))
        raise

def Analise_Alteracoes(**projetos):
    try:
        dt = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"); print ("{} - Conectando ao banco para analise das alterações ...".format(dt))

        BASE_DIR = os.path.dirname(os.path.abspath(__file__))
        db_path = os.path.join(BASE_DIR, "Projetos.db")

        conn = sqlite3.connect(db_path)
        cursor = conn.cursor()

        Nome_do_Projeto = projetos['Nome_do_Projeto']

        dt = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"); print("Analisando alterações Formula Fase do Projeto: {}".format(Nome_do_Projeto))

        id_arquivo = projetos ['id_arquivo']

        cursor.execute("""
        SELECT id, Formula_Fase,Formula_Status_Projeto,Ini_desenv,Term_desenv,Ini_Teste_Integrado,Term_Teste_Integrado,Ini_hml, Fim_hml,Problema_Risco, SubCausa, Plano_Acao, causa FROM projetos
        WHERE Nome_Projeto = ?
        """, (Nome_do_Projeto, ))

        for id_projeto, Formula_Fase,Formula_Status_Projeto,Ini_desenv,Term_desenv,Ini_Teste_Integrado,Term_Teste_Integrado,Ini_hml, Fim_hml,Problema_Risco, SubCausa, Plano_Acao, causa in cursor.fetchall():
            # bloco das comparações se ocorreu mudança nas informações da planilha
            if (Formula_Fase != projetos ['Formula_Fase']):
                CodEvento = 200
                descricao = "Formula Fase Alterada" #evento
                valor_anterior = Formula_Fase
                novo_valor = projetos ['Formula_Fase']
                #Guardar mudança encontrada
                Insert_Historico(id_projeto, id_arquivo, CodEvento, descricao, valor_anterior, novo_valor)
                update_projeto (id_projeto, id_arquivo, CodEvento, novo_valor)
            
            elif (Formula_Status_Projeto != projetos ['Formula_Status_Projeto']):
                CodEvento = 300
                descricao = "Formula status foi Alterada" #evento
                valor_anterior = Formula_Status_Projeto
                novo_valor = projetos ['Formula_Status_Projeto']
                #Guardar mudança encontrada
                Insert_Historico(id_projeto, id_arquivo, CodEvento, descricao, valor_anterior, novo_valor)
                update_projeto (id_projeto, id_arquivo, CodEvento, novo_valor)

            elif (Ini_desenv != projetos ['Inicio_desenv']):
                CodEvento = 400
                descricao = "Data Inicio do Desenvolvimento foi Alterada" #evento
                valor_anterior = Ini_desenv
                novo_valor = projetos ['Inicio_desenv']
                #Guardar mudança encontrada
                Insert_Historico(id_projeto, id_arquivo, CodEvento, descricao, valor_anterior, novo_valor)
                update_projeto (id_projeto, id_arquivo, CodEvento, novo_valor)
                
            elif (Term_desenv != projetos ['Termino_Desenv']):
                CodEvento = 450
                descricao = "Data Término do Desenvolvimento foi Alterada" #evento
                valor_anterior = Term_desenv
                novo_valor = projetos ['Termino_Desenv']
                #Guardar mudança encontrada
                Insert_Historico(id_projeto, id_arquivo, CodEvento, descricao, valor_anterior, novo_valor)
                update_projeto (id_projeto, id_arquivo, CodEvento, novo_valor)

            elif (Ini_Teste_Integrado != projetos ['Inicio_Testes_Integrados']):
                CodEvento = 500
                descricao = "Data Inicio do Teste Integrado foi Alterada" #evento
                valor_anterior = Ini_Teste_Integrado
                novo_valor = projetos ['Inicio_Testes_Integrados']
                #Guardar mudança encontrada
                Insert_Historico(id_projeto, id_arquivo, CodEvento, descricao, valor_anterior, novo_valor)
                update_projeto (id_projeto, id_arquivo, CodEvento, novo_valor)

            elif (Term_Teste_Integrado != projetos ['Termino_Testes_Integrados']):
                CodEvento = 550
                descricao = "Data Término do Teste Integrado foi Alterada" #evento
                valor_anterior = Term_Teste_Integrado
                novo_valor = projetos ['Termino_Testes_Integrados']
                #Guardar mudança encontrada
                Insert_Historico(id_projeto, id_arquivo, CodEvento, descricao, valor_anterior, novo_valor)
                update_projeto (id_projeto, id_arquivo, CodEvento, novo_valor)

            elif (Ini_hml != projetos ['Inicio_da_HML']):
                CodEvento = 600
                descricao = "Data Inicio da Homologação foi Alterada" #evento
                valor_anterior = Ini_hml
                novo_valor = projetos ['Inicio_da_HML']
                #Guardar mudança encontrada
                Insert_Historico(id_projeto, id_arquivo, CodEvento, descricao, valor_anterior, novo_valor)
                update_projeto (id_projeto, id_arquivo, CodEvento, novo_valor)

            elif (Fim_hml != projetos ['Termino_da_HML']):
                CodEvento = 650
                descricao = "Data Término da Homologação foi Alterada" #evento
                valor_anterior = Fim_hml
                novo_valor = projetos ['Termino_da_HML']
                #Guardar mudança encontrada
                Insert_Historico(id_projeto, id_arquivo, CodEvento, descricao, valor_anterior, novo_valor)
                update_projeto (id_projeto, id_arquivo, CodEvento, novo_valor)

            elif (Problema_Risco != projetos ['Problemas_Riscos']):
                CodEvento = 700
                descricao = "Problema e Risco foi Alterado" #evento
                valor_anterior = Problema_Risco
                novo_valor = projetos ['Problemas_Riscos']
                #Guardar mudança encontrada
                Insert_Historico(id_projeto, id_arquivo, CodEvento, descricao, valor_anterior, novo_valor)
                update_projeto (id_projeto, id_arquivo, CodEvento, novo_valor)

            elif (SubCausa != projetos ['SubCausa_Capacity']):
                CodEvento = 800
                descricao = "Sub causa foi Alterada" #evento
                valor_anterior = SubCausa
                novo_valor = projetos ['SubCausa_Capacity']
                #Guardar mudança encontrada
                Insert_Historico(id_projeto, id_arquivo, CodEvento, descricao, valor_anterior, novo_valor)
                update_projeto (id_projeto, id_arquivo, CodEvento, novo_valor)

            elif (Plano_Acao != projetos ['Plano_de_Acao']):
                CodEvento = 900
                descricao = "Plano de ação foi Alterado" #evento
                valor_anterior = Plano_Acao
                novo_valor = projetos ['Plano_de_Acao']
                #Guardar mudança encontrada
                Insert_Historico(id_projeto, id_arquivo, CodEvento, descricao, valor_anterior, novo_valor)
                update_projeto (id_projeto, id_arquivo, CodEvento, novo_valor)

            elif (causa != projetos ['Causa']):
                CodEvento = 1000
                descricao = "Causa foi Alterado" #evento
                valor_anterior = causa
                novo_valor = projetos ['Causa']
                #Guardar mudança encontrada
                Insert_Historico(id_projeto, id_arquivo, CodEvento, descricao, valor_anterior, novo_valor)
                update_projeto (id_projeto, id_arquivo, CodEvento, novo_valor)
    except Exception as e:
        print ("Erro gerado: \n {}".format(str(e)))
        raise
def update_projeto (id_projeto, id_arquivo, CodEvento, novo_valor):
    try:
        dt = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"); print ("{} - Conectando ao banco e inserindo historico ...".format(dt))
        BASE_DIR = os.path.dirname(os.path.abspath(__file__))
        db_path = os.path.join(BASE_DIR, "Projetos.db")
        
        conn = sqlite3.connect(db_path)
        cursor = conn.cursor()

        if (CodEvento == 200):
            #update formula fase
            cursor.execute("""
            update projetos
            set Formula_Fase = ?
            where id_projeto = ?
            """, (novo_valor,id_projeto))

        elif (CodEvento == 300):
            #update Formula_Status_Projeto
            cursor.execute("""
            update projetos
            set Formula_Status_Projeto = ?
            where id_projeto = ?
            """, (novo_valor,id_projeto))

        elif (CodEvento == 400):
            #update Ini_desenv
            cursor.execute("""
            update projetos
            set Ini_desenv = ?
            where id_projeto = ?
            """, (novo_valor,id_projeto))

        elif (CodEvento == 450):
            #update Term_desenv
            cursor.execute("""
            update projetos
            set Term_desenv = ?
            where id_projeto = ?
            """, (novo_valor,id_projeto))

        elif (CodEvento == 500):
            #update Ini_Teste_Integrado
            cursor.execute("""
            update projetos
            set Ini_Teste_Integrado = ?
            where id_projeto = ?
            """, (novo_valor,id_projeto))

        elif (CodEvento == 550):
            #update Term_Teste_Integrado
            cursor.execute("""
            update projetos
            set Term_Teste_Integrado = ?
            where id_projeto = ?
            """, (novo_valor,id_projeto))

        elif (CodEvento == 600):
            #update Ini_hml
            cursor.execute("""
            update projetos
            set Ini_hml = ?
            where id_projeto = ?
            """, (novo_valor,id_projeto))

        elif (CodEvento == 650):
            #update Fim_hml
            cursor.execute("""
            update projetos
            set Fim_hml = ?
            where id_projeto = ?
            """, (novo_valor,id_projeto))

        elif (CodEvento == 700):
            #update Problema_Risco
            cursor.execute("""
            update projetos
            set Problema_Risco = ?
            where id_projeto = ?
            """, (novo_valor,id_projeto))

        elif (CodEvento == 800):
            #update SubCausa
            cursor.execute("""
            update projetos
            set SubCausa = ?
            where id_projeto = ?
            """, (novo_valor,id_projeto))

        elif (CodEvento == 900):
            #update Plano_Acao
            cursor.execute("""
            update projetos
            set Plano_Acao = ?
            where id_projeto = ?
            """, (novo_valor,id_projeto))

        elif (CodEvento == 1000):
            #update causa
            cursor.execute("""
            update projetos
            set causa = ?
            where id_projeto = ?
            """, (novo_valor,id_projeto))

        #atualiza o arquivo
        cursor.execute("""
        update projetos
        set id_arquivo = ?
        where id_projeto = ?
        """, (id_arquivo,id_projeto))
        
        conn.commit()
        #dt = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"); print ("{} - Projeto atualizado...".format(dt))
        conn.close()
    
    except Exception as e:
        raise
    finally:
        if (conn.close()):
            conn.close()

def Insert_Historico(id_projeto, id_arquivo, CodEvento, descricao, valor_anterior, novo_valor):
    try:
        dt = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"); print ("{} - Conectando ao banco e inserindo historico ...".format(dt))
        

        BASE_DIR = os.path.dirname(os.path.abspath(__file__))
        db_path = os.path.join(BASE_DIR, "Projetos.db")
        
        conn = sqlite3.connect(db_path)
        cursor = conn.cursor()

        cursor.execute("""
        INSERT INTO historico (id_projeto, id_arquivo, cod_evento, evento, valor_anterior, novo_valor)
        VALUES (?, ?, ?, ?, ?, ?)
        """, (id_projeto,id_arquivo,CodEvento,descricao,valor_anterior,novo_valor,))

        conn.commit()

        dt = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"); print ("{} - Histórico inserido com sucesso...".format(dt))

        conn.close()
    except Exception as e:
        print ("Erro gerado: \n {}".format(str(e)))
        raise


def insert_new_proj(**projetos):
    try:
        dt = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"); print ("{} - Conectando ao banco e inserindo linha ...".format(dt))
        

        BASE_DIR = os.path.dirname(os.path.abspath(__file__))
        db_path = os.path.join(BASE_DIR, "Projetos.db")
        
        conn = sqlite3.connect(db_path)
        cursor = conn.cursor()
        
        Nome_do_Projeto = projetos['Nome_do_Projeto']
        VP = projetos['VP']
        Formula_Fase= projetos['Formula_Fase']
        A_N= projetos['A_N']
        GP = projetos['GP']
        LT = projetos['LT']
        Lider_de_Testes = projetos ['Lider_de_Testes']
        Gerente_de_Desenvolvimento = projetos['Gerente_de_Desenvolvimento']
        Formula_Status_Projeto = projetos ['Formula_Status_Projeto']
        Descricao = projetos ['Descricao']
        Nome_Arquivo = ""
        Inicio_desenv = projetos ['Inicio_desenv']
        Termino_Desenv = projetos ['Termino_Desenv']
        Complitude1 = projetos ['Complitude1']
        Inicio_Testes_Integrados = projetos ['Inicio_Testes_Integrados']
        Termino_Testes_Integrados = projetos ['Termino_Testes_Integrados']
        Complitude2 = projetos ['Complitude2']
        Inicio_da_HML = projetos ['Inicio_da_HML']
        Termino_da_HML = projetos ['Termino_da_HML']
        Complitude3 = projetos ['Complitude3']
        Problemas_Riscos = projetos ['Problemas_Riscos']
        SubCausa_Capacity = projetos ['SubCausa_Capacity']
        Plano_de_Acao = projetos ['Plano_de_Acao']
        Causa = projetos ['Causa']
        id_arquivo = projetos ['id_arquivo']

        print("Adicionando o projeto: {}".format(Nome_do_Projeto))

        cursor.execute("""
        INSERT INTO projetos (Nome_Projeto,VP,Formula_Fase, an, gp, LT, Lider_Teste, Gerente_Desenv, Formula_Status_Projeto, Descricao, Nome_Arquivo, Ini_desenv, Term_desenv, Completude1, Ini_Teste_Integrado, Term_Teste_Integrado, Completude2,Ini_hml, Fim_hml, Completude3, Problema_Risco, SubCausa, Plano_Acao, causa, id_arquivo)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?,?, ?, ?, ?, ?, ?, ?, ?)
        """, (Nome_do_Projeto, VP, Formula_Fase, A_N, GP, LT, Lider_de_Testes, Gerente_de_Desenvolvimento, Formula_Status_Projeto, Descricao, Nome_Arquivo, Inicio_desenv, Termino_Desenv, Complitude1, Inicio_Testes_Integrados, Termino_Testes_Integrados, Complitude2, Inicio_da_HML, Termino_da_HML, Complitude3, Problemas_Riscos, SubCausa_Capacity, Plano_de_Acao, Causa, id_arquivo,))

        conn.commit()

        dt = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"); print ("{} - Novos Projetos Inseridos com sucesso...".format(dt))

        conn.close()

    except Exception as e:
        print ("Erro gerado: \n {}".format(str(e)))
        #raise
        

def ler_arquivo(wb):
    _wb = wb
    dt = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    print("() - {}".format(dt,_wb.sheetnames))

def creation_date(path_to_file):
    """
    Try to get the date that a file was created, falling back to when it was
    last modified if that isn't possible.
    See http://stackoverflow.com/a/39501288/1709587 for explanation.
    """
    if platform.system() == 'Windows':
        return os.path.getctime(path_to_file)
    else:
        stat = os.stat(path_to_file)
        try:
            return stat.st_birthtime
        except AttributeError:
            # We're probably on Linux. No easy way to get creation dates here,
            # so we'll settle for when its content was last modified.
            return stat.st_mtime
if __name__ == "__main__":
	main()

