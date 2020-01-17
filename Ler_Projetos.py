# -*- coding: utf-8 -*
import datetime
import sys
import glob, os
import sqlite3
from openpyxl import load_workbook

def main():
    lerdir01()
    #lerdir03()

def lerdir01():
    colNome_do_Projeto = 1
    colVP = 3
    colFormula_Fase = 95
    colA_N = 45
    colGP = 46
    colLT = 47
    colLider_de_Testes = 48
    colGerente_de_Desenvolvimento = 49
    colFormula_Status_Projeto = 107
    colDescricao = 44
    colInicio_desenv = 29
    colTermino_Desenv = 30
    colComplitude1 = 31
    colInicio_Testes_Integrados = 32
    colTermino_Testes_Integrados = 33
    colComplitude2 = 34
    colInicio_da_HML = 35
    colTermino_da_HML = 36
    colComplitude3 = 37
    colProblemas_Riscos = 62
    colPlano_de_Acao = 63
    colCausa = 64
    colSubCausa_Capacity = 65
    
    colLeonardo = "Leonardo Augusto Mendes Leandro"
    colJuliana = "Juliana Alves Castro Perez"
    
    os.chdir("pendentes")
    #for file in glob.glob("*.xlsm"):
    #print(file)
    dict = {}
    for i in os.listdir(os.getcwd()):
        #print(i)
        dict.clear();
        print (os.path.join(os.getcwd(), i))
        if os.path.isfile(i):
            if i.endswith(".xls") or i.endswith(".xlsx") or i.endswith(".xlsm"):
                wb = load_workbook(os.path.join(os.getcwd(), i))
                ws_Projetos = wb['Projetos']
                total_linhas = ws_Projetos.max_row
                print(total_linhas)
                linha = 4
                dict ['Nome_do_Projeto'] = ws_Projetos.cell(row=linha, column=colNome_do_Projeto).value
                dict ['VP'] = ws_Projetos.cell(row=linha, column=colVP).value
                dict ['Formula_Fase'] = ws_Projetos.cell(row=linha, column=colFormula_Fase).value
                dict ['A_N'] = ws_Projetos.cell(row=linha, column=colA_N).value
                dict ['VP'] = ws_Projetos.cell(row=linha, column=colVP).value
                dict ['GP'] = ws_Projetos.cell(row=linha, column=colGP).value
                dict ['LT'] = ws_Projetos.cell(row=linha, column=colLT).value
                dict ['Lider_de_Testes'] = ws_Projetos.cell(row=linha, column=colLider_de_Testes).value
                dict ['Gerente_de_Desenvolvimento'] = ws_Projetos.cell(row=linha, column=colGerente_de_Desenvolvimento).value
                dict ['Formula_Status_Projeto'] = ws_Projetos.cell(row=linha, column=colFormula_Status_Projeto).value
                dict ['Descricao'] = ws_Projetos.cell(row=linha, column=colDescricao).value
                dict ['Inicio_desenv'] = ws_Projetos.cell(row=linha, column=colInicio_desenv).value
                dict ['Termino_Desenv'] = ws_Projetos.cell(row=linha, column=colTermino_Desenv).value
                dict ['Complitude1'] = ws_Projetos.cell(row=linha, column=colComplitude1).value
                dict ['Inicio_Testes_Integrados'] = ws_Projetos.cell(row=linha, column=colInicio_Testes_Integrados).value
                dict ['Termino_Testes_Integrados'] = ws_Projetos.cell(row=linha, column=colTermino_Testes_Integrados).value
                dict ['Complitude2'] = ws_Projetos.cell(row=linha, column=colComplitude2).value
                dict ['Inicio_da_HML'] = ws_Projetos.cell(row=linha, column=colInicio_da_HML).value
                dict ['Termino_da_HML'] = ws_Projetos.cell(row=linha, column=colTermino_da_HML).value
                dict ['Complitude3'] = ws_Projetos.cell(row=linha, column=colComplitude3).value
                dict ['Problemas_Riscos'] = ws_Projetos.cell(row=linha, column=colProblemas_Riscos).value
                dict ['Plano_de_Acao'] = ws_Projetos.cell(row=linha, column=colPlano_de_Acao).value
                dict ['Causa'] = ws_Projetos.cell(row=linha, column=colCausa).value
                dict ['SubCausa_Capacity'] = ws_Projetos.cell(row=linha, column=colSubCausa_Capacity).value
                dict ['Causa'] = ws_Projetos.cell(row=linha, column=colCausa).value
                lerdict(dict)

def lerdict(dict):
    dt = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"); print ("{} - MONTANDO LISTA ... ...".format(dt))
    #print ("A_N: {}".format(dict['A_N']))
    linha = []
    linha.append(dict ['Nome_do_Projeto'])
    linha.append(dict ['VP'])
    linha.append(dict ['Formula_Fase'])
    linha.append(dict ['A_N'])
    linha.append(dict ['GP'])
    linha.append(dict ['Lider_de_Testes'])
    linha.append(dict ['Gerente_de_Desenvolvimento'])
    linha.append(dict ['Formula_Status_Projeto'])
    linha.append(dict ['Descricao'])
    linha.append("") #Nome_arquivo
    linha.append(dict ['Inicio_desenv'])
    linha.append(dict ['Termino_Desenv'])
    linha.append(dict ['Complitude1'])
    linha.append(dict ['Inicio_Testes_Integrados'])
    linha.append(dict ['Termino_Testes_Integrados'])
    linha.append(dict ['Complitude2'])
    linha.append(dict ['Inicio_da_HML'])
    linha.append(dict ['Termino_da_HML'])
    linha.append(dict ['Complitude3'])
    linha.append(dict ['Problemas_Riscos'])
    linha.append(dict ['SubCausa_Capacity'])
    linha.append(dict ['Plano_de_Acao'])

    insert_db(linha)

def insert_db(linha):
    dt = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"); print ("{} - Conectando ao banco e inserindo linha ...".format(dt))

    conn = sqlite3.connect('Projetos.db')
    cursor = conn.cursor()
    # inserindo dados na tabela
    
    cursor.executemany("""
    INSERT INTO projetos (Nome_Projeto, VP, Formula_Fase, an, gp, Lider_Teste, Gerente_Desenv, Formula_Status_Projeto, Descricao, Nome_Arquivo, Ini_desenv, Term_desenv, Completude1, Ini_Teste_Integrado, Term_Teste_Integrado, Completude2, Ini_hml, Fim_hml, Completude3, Problema_Risco, SubCausa, Plano_Acao)
    VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
    """, linha)

    conn.commit()

    dt = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"); print ("{} - DADOS INSERIDOS COM SUCESSO ...".format(dt))

    conn.close()

def lerdir03():
    os.chdir("pendentes")
    files = glob.glob("*.xlsm")
    files.sort(key=os.path.getctime)
    #files.sort(key=os.path.getmtime)
    print("\n".join(files))
    files1 = ("\n".join(files))
    for file in files1:
        print(file)
        

def lerdir04():
    for root, dirs, files in os.walk("pendentes"):
        for file in files:
            if file.endswith(".xlsm"):
                 print(os.path.join(root, file))

def ler_arquivo(wb):
    _wb = wb
    dt = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    print("() - {}".format(dt,_wb.sheetnames))


if __name__ == "__main__":
	main()

